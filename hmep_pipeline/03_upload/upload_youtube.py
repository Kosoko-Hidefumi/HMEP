"""
③ YouTube アップロード: videos_renamed/ の動画を YouTube Data API v3 へ送る。

初回のみブラウザ認証（OAuth）。トークンは credentials/token.json に保存する。

使用例（hmep_pipeline/03_upload から）:
  python upload_youtube.py --dry-run
  python upload_youtube.py
"""

from __future__ import annotations

import argparse
import csv
import logging
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import yaml
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload
from openpyxl import load_workbook

PIPELINE_ROOT = Path(__file__).resolve().parent.parent

# OAuth: upload + プレイリスト登録
SCOPES = ("https://www.googleapis.com/auth/youtube",)

UPLOAD_COLS: tuple[str, ...] = (
    "youtube_video_id",
    "youtube_url",
    "upload_status",
    "uploaded_at",
)

UPLOAD_LOG_FIELDS = (
    "uploaded_at",
    "local_file",
    "video_id",
    "youtube_url",
    "status",
    "error_message",
)


def load_config(path: Path) -> dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def resolve_path(root: Path, rel: str) -> Path:
    p = Path(rel)
    if p.is_absolute():
        return p.resolve()
    return (root / rel).resolve()


def to_date(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    try:
        s = str(val)[:10]
        return date.fromisoformat(s)
    except (ValueError, TypeError):
        return None


def format_lecture_date_for_template(val: Any) -> str:
    d = to_date(val)
    if d is None:
        return str(val or "").strip() or "（不明）"
    return d.isoformat()


def load_sheet(xlsx: Path) -> tuple[list[str], list[dict[str, Any]]]:
    if not xlsx.is_file():
        return [], []
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    if "lectures" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"{xlsx} に lectures シートがありません")
    ws = wb["lectures"]
    headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        if all(v is None for v in row):
            continue
        d: dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            d[h] = row[i] if i < len(row) else None
        rows.append(d)
    wb.close()
    return headers, rows


def save_lectures_sheet(
    xlsx: Path,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> None:
    wb = load_workbook(xlsx)
    if "lectures" not in wb.sheetnames:
        wb.close()
        raise ValueError("lectures sheet missing")
    ws = wb["lectures"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h or None)
    for ri, rd in enumerate(rows, start=2):
        for c, h in enumerate(headers, start=1):
            if not h:
                continue
            ws.cell(row=ri, column=c, value=rd.get(h))
    wb.save(xlsx)
    wb.close()


def setup_logging(log_dir: Path) -> None:
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / f"upload_{datetime.now().strftime('%Y-%m-%d')}.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )


def load_description_template(path: Path) -> str:
    if not path.is_file():
        raise FileNotFoundError(f"説明テンプレートがありません: {path}")
    return path.read_text(encoding="utf-8")


def bare_speaker_name(val: Any) -> str:
    """テンプレートの「{speaker_name} 先生」用に、末尾の「先生」を除いた氏名を返す。"""
    s = str(val or "").strip()
    if s.endswith("先生"):
        return s[: -len("先生")].strip()
    return s


def sanitize_youtube_description(text: str) -> str:
    """YouTube が snippet.description で拒否しがちな表記を弱める（例: mailto の角括弧付き）。"""
    # Outlook 由来の "email <mailto:email>" は invalidDescription になることがある
    text = re.sub(r"\s*<mailto:[^>]+>", "", text)
    return text


def build_description(row: dict[str, Any], template: str) -> str:
    def g(key: str) -> str:
        v = row.get(key)
        if v is None:
            return ""
        return str(v).strip()

    ctx = {
        "lecture_date": format_lecture_date_for_template(row.get("lecture_date")),
        "target_grade": g("target_grade") or "—",
        "title": g("title"),
        "speaker_name": bare_speaker_name(row.get("speaker_name")),
        "speaker_affiliation": g("speaker_affiliation"),
        "description": g("description"),
        "biography": g("biography"),
    }

    out = sanitize_youtube_description(template.format(**ctx))
    if len(out) > 5000:
        return out[:4997] + "..."
    return out


def mime_for_video(path: Path) -> str:
    m = {
        ".mp4": "video/mp4",
        ".mov": "video/quicktime",
        ".mkv": "video/x-matroska",
        ".avi": "video/x-msvideo",
        ".m4v": "video/x-m4v",
    }
    return m.get(path.suffix.lower(), "application/octet-stream")


def truncate_title(stem: str, max_len: int = 100) -> str:
    s = stem.strip()
    if len(s) <= max_len:
        return s
    return s[: max_len - 1] + "…"


def build_tags(row: dict[str, Any]) -> list[str]:
    ld = to_date(row.get("lecture_date"))
    year = str(ld.year) if ld else ""
    speaker = str(row.get("speaker_name") or "").strip()
    for suf in ("先生", " 先生"):
        if speaker.endswith(suf):
            speaker = speaker[: -len(suf)].strip()
            break
    tags = ["HMEP"]
    if year:
        tags.append(year)
    if speaker:
        tags.append(speaker[:80])
    return tags[:30]


def run_auth_only(config_path: Path) -> int:
    """OAuth のみ実行し token.json を保存する（lectures.xlsx は不要）。"""
    cfg = load_config(config_path)
    root = config_path.parent
    yt = cfg.get("youtube") or {}
    paths = cfg.get("paths") or {}
    logs_rel = paths.get("logs_dir") or "logs"

    client_secrets = resolve_path(root, yt.get("client_secrets_file") or "credentials/client_secrets.json")
    token_path = resolve_path(root, yt.get("token_file") or "credentials/token.json")
    log_dir = resolve_path(root, logs_rel)

    setup_logging(log_dir)
    try:
        get_youtube_service(client_secrets, token_path)
    except FileNotFoundError as e:
        logging.error("%s", e)
        return 2
    logging.info("OAuth 完了。トークン: %s", token_path)
    return 0


def get_youtube_service(client_secrets: Path, token_path: Path):
    creds: Optional[Credentials] = None
    if token_path.is_file():
        creds = Credentials.from_authorized_user_file(str(token_path), list(SCOPES))
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not client_secrets.is_file():
                raise FileNotFoundError(
                    f"OAuth クライアントファイルがありません: {client_secrets}\n"
                    "Google Cloud で YouTube Data API v3 を有効化し、"
                    "デスクトップ用 OAuth クライアントの JSON を配置してください。"
                )
            flow = InstalledAppFlow.from_client_secrets_file(str(client_secrets), list(SCOPES))
            creds = flow.run_local_server(port=0)
        token_path.parent.mkdir(parents=True, exist_ok=True)
        with open(token_path, "w", encoding="utf-8") as token:
            token.write(creds.to_json())
    return build("youtube", "v3", credentials=creds)


def upload_one_video(
    service: Any,
    file_path: Path,
    title: str,
    description: str,
    tags: list[str],
    privacy_status: str,
    category_id: str,
) -> str:
    body: dict[str, Any] = {
        "snippet": {
            "title": title,
            "description": description,
            "tags": tags,
            "categoryId": category_id,
            "defaultLanguage": "ja",
            "defaultAudioLanguage": "ja",
        },
        "status": {
            "privacyStatus": privacy_status,
            "selfDeclaredMadeForKids": False,
        },
    }
    media = MediaFileUpload(
        str(file_path),
        mimetype=mime_for_video(file_path),
        resumable=True,
        chunksize=1024 * 1024,
    )
    req = service.videos().insert(part="snippet,status", body=body, media_body=media)
    response = None
    while response is None:
        status, response = req.next_chunk()
        if status:
            p = status.progress()
            if p is not None:
                logging.info("  ... アップロード進捗 %.0f%%", p * 100)
    vid = (response or {}).get("id")
    if not vid:
        raise RuntimeError("YouTube API が video id を返しませんでした")
    return str(vid)


def add_to_playlist(service: Any, playlist_id: str, video_id: str) -> None:
    service.playlistItems().insert(
        part="snippet",
        body={
            "snippet": {
                "playlistId": playlist_id,
                "resourceId": {"kind": "youtube#video", "videoId": video_id},
            }
        },
    ).execute()


def load_uploaded_local_files(log_path: Path) -> set[str]:
    if not log_path.is_file():
        return set()
    out: set[str] = set()
    with open(log_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if (row.get("status") or "").strip().lower() != "ok":
                continue
            lf = (row.get("local_file") or "").strip()
            if lf:
                out.add(lf)
    return out


def append_upload_log(log_path: Path, row: dict[str, str]) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    new_file = not log_path.is_file()
    with open(log_path, "a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(UPLOAD_LOG_FIELDS), extrasaction="ignore")
        if new_file:
            w.writeheader()
        w.writerow(row)
        f.flush()


def run_upload(
    config_path: Path,
    dry_run: bool,
    force: bool,
    limit_override: Optional[int],
) -> int:
    cfg = load_config(config_path)
    root = config_path.parent

    r_cfg = cfg.get("rename") or {}
    outlook = cfg.get("outlook") or {}
    date_from_s = r_cfg.get("date_from") or outlook.get("lecture_date_from") or "2025-01-01"
    date_to_s = r_cfg.get("date_to") or outlook.get("lecture_date_to") or "2026-05-31"
    start_d = date.fromisoformat(str(date_from_s)[:10])
    end_d = date.fromisoformat(str(date_to_s)[:10])

    paths = cfg.get("paths") or {}
    yt = cfg.get("youtube") or {}

    out_rel = paths.get("videos_renamed_dir") or "videos_renamed"
    xlsx_rel = paths.get("lectures_xlsx") or "data/lectures.xlsx"
    log_rel = paths.get("upload_log_csv") or "data/upload_log.csv"
    logs_rel = paths.get("logs_dir") or "logs"

    dest_dir = resolve_path(root, out_rel)
    xlsx_path = resolve_path(root, xlsx_rel)
    upload_log_path = resolve_path(root, log_rel)
    log_dir = resolve_path(root, logs_rel)

    client_secrets = resolve_path(root, yt.get("client_secrets_file") or "credentials/client_secrets.json")
    token_path = resolve_path(root, yt.get("token_file") or "credentials/token.json")
    template_path = resolve_path(root, yt.get("description_template_file") or "03_upload/description_template.txt")

    privacy = str(yt.get("privacy_status") or "unlisted")
    category_id = str(yt.get("category_id") or "27")
    playlist_id = str(yt.get("playlist_id") or "").strip()
    max_uploads = int(yt.get("max_uploads_per_run") or 0)
    if limit_override is not None:
        max_uploads = limit_override

    setup_logging(log_dir)

    if not xlsx_path.is_file():
        logging.error("lectures.xlsx がありません: %s", xlsx_path)
        return 2

    try:
        tpl = load_description_template(template_path)
    except FileNotFoundError as e:
        logging.error("%s", e)
        return 2

    headers, rows = load_sheet(xlsx_path)
    for c in UPLOAD_COLS:
        if c not in headers:
            headers.append(c)

    already_logged = load_uploaded_local_files(upload_log_path)

    service = None
    if not dry_run:
        try:
            service = get_youtube_service(client_secrets, token_path)
        except FileNotFoundError as e:
            logging.error("%s", e)
            return 2

    n_ok = 0
    n_fail = 0
    n_skip = 0
    n_dry = 0
    attempts_this_run = 0  # dry-run 含め、この実行で処理する件数（上限との比較用）

    for row in rows:
        ld = to_date(row.get("lecture_date"))
        if ld is None or ld < start_d or ld > end_d:
            continue

        fname = str(row.get("renamed_video_file") or "").strip()
        if not fname:
            continue

        video_path = (dest_dir / fname).resolve()
        base_title = Path(fname).stem

        if not video_path.is_file():
            if not dry_run:
                row["upload_status"] = "skipped_file_missing"
            logging.warning("ファイルなし: %s", video_path)
            n_skip += 1
            continue

        existing_url = str(row.get("youtube_url") or "").strip()
        existing_id = str(row.get("youtube_video_id") or "").strip()
        if not force and (existing_url or existing_id):
            if not dry_run:
                row.setdefault("upload_status", "skipped_already_in_sheet")
            n_skip += 1
            continue

        if not force and fname in already_logged:
            if not dry_run:
                row.setdefault("upload_status", "skipped_in_upload_log")
            n_skip += 1
            continue

        if max_uploads > 0 and attempts_this_run >= max_uploads:
            if not dry_run:
                row.setdefault("upload_status", "skipped_quota_limit")
            continue

        attempts_this_run += 1

        title = truncate_title(base_title)
        description = build_description(row, tpl)
        tags = build_tags(row)
        now_s = datetime.now().isoformat(timespec="seconds")

        if dry_run:
            logging.info("[dry-run] upload: %s | title=%s", fname, title[:60])
            n_dry += 1
            continue

        assert service is not None
        try:
            vid = upload_one_video(service, video_path, title, description, tags, privacy, category_id)
            url = f"https://www.youtube.com/watch?v={vid}"
            if playlist_id:
                try:
                    add_to_playlist(service, playlist_id, vid)
                except HttpError as e:
                    logging.error("プレイリスト登録に失敗（動画はアップ済み）: %s", e)

            row["youtube_video_id"] = vid
            row["youtube_url"] = url
            row["upload_status"] = "ok"
            row["uploaded_at"] = now_s

            append_upload_log(
                upload_log_path,
                {
                    "uploaded_at": now_s,
                    "local_file": fname,
                    "video_id": vid,
                    "youtube_url": url,
                    "status": "ok",
                    "error_message": "",
                },
            )
            already_logged.add(fname)
            n_ok += 1
            logging.info("アップロード済み: %s -> %s", fname, url)
        except HttpError as e:
            err = str(e)
            row["upload_status"] = "error"
            n_fail += 1
            append_upload_log(
                upload_log_path,
                {
                    "uploaded_at": now_s,
                    "local_file": fname,
                    "video_id": "",
                    "youtube_url": "",
                    "status": "error",
                    "error_message": err[:2000],
                },
            )
            logging.error("アップロード失敗 %s: %s", fname, err)
        except Exception as e:
            err = str(e)
            row["upload_status"] = "error"
            n_fail += 1
            append_upload_log(
                upload_log_path,
                {
                    "uploaded_at": now_s,
                    "local_file": fname,
                    "video_id": "",
                    "youtube_url": "",
                    "status": "error",
                    "error_message": err[:2000],
                },
            )
            logging.error("アップロード失敗 %s: %s", fname, err)

    if not dry_run:
        save_lectures_sheet(xlsx_path, headers, rows)

    if dry_run:
        logging.info("dry-run 終了: アップロード対象 %d 件（台帳は未保存）", n_dry)
    else:
        logging.info("完了: ok=%d fail=%d skip=%d", n_ok, n_fail, n_skip)
    return 1 if (not dry_run and n_fail > 0) else 0


def main(argv: Optional[list[str]] = None) -> int:
    p = argparse.ArgumentParser(description="HMEP YouTube upload (videos_renamed → YouTube)")
    p.add_argument(
        "--config",
        type=Path,
        default=PIPELINE_ROOT / "config.yaml",
        help="config.yaml のパス（既定: hmep_pipeline/config.yaml）",
    )
    p.add_argument("--dry-run", action="store_true", help="API を呼ばず対象のみ表示")
    p.add_argument("--force", action="store_true", help="台帳・ログに既存があっても再アップロード")
    p.add_argument("--limit", type=int, default=None, help="この実行でアップロードする最大件数（クォータ対策）")
    p.add_argument(
        "--auth-only",
        action="store_true",
        help="ブラウザ認証のみ行い token.json を保存して終了（台帳なしで初回セットアップ可）",
    )
    args = p.parse_args(argv)
    cfg_path = Path(args.config).resolve()
    if args.auth_only:
        return run_auth_only(cfg_path)
    return run_upload(cfg_path, dry_run=args.dry_run, force=args.force, limit_override=args.limit)


if __name__ == "__main__":
    sys.exit(main())

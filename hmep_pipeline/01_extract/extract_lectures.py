"""
HMEP ① Outlook メールからレクチャー情報を抽出し data/lectures.xlsx を更新する。

起動前に Outlook デスクトップを起動しておくこと。

使用例:
  python extract_lectures.py
  python extract_lectures.py --config ..\\config.yaml
  python extract_lectures.py --fixture sample_mail.json   # Outlook なしでパース検証
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator, Optional

import yaml

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
except ImportError as e:
    raise SystemExit("openpyxl is required. pip install -r requirements.txt") from e

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PIPELINE_ROOT = SCRIPT_DIR.parent


def load_config(path: Path) -> dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def resolve_under_root(root: Path, relative: str) -> Path:
    p = (root / relative).resolve()
    return p


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

DEFAULT_SUBJECT_DATE_PATTERNS: tuple[str, ...] = (
    r"(?P<y>\d{4})\s*[/／年]\s*(?P<m>\d{1,2})\s*[/／月]\s*(?P<d>\d{1,2})",
    r"(?P<y>\d{4})\s*[/／]\s*(?P<m>\d{1,2})\s*[/／]\s*(?P<d>\d{1,2})",
    r"(?P<y>\d{4})\s*年\s*(?P<m>\d{1,2})\s*月\s*(?P<d>\d{1,2})\s*日",
)

MARKERS = (
    ("target_grade", "【全学年対象】", "【タイトル】"),
    ("title", "【タイトル】", "【講師】"),
    ("speaker_block", "【講師】", "【講義紹介文】"),
    ("description", "【講義紹介文】", "【講師略歴】"),
    ("biography", "【講師略歴】", None),
)

SPEAKER_END_MARKERS = ("【講義紹介文】", "【講義紹介】")
DESC_START_MARKERS = ("【講義紹介文】", "【講義紹介】")
BIO_START_MARKERS = ("【講師略歴】", "【講師紹介】", "【講師紹介文】")


def _slice_between(text: str, start_m: str, end_m: Optional[str]) -> Optional[str]:
    i = text.find(start_m)
    if i == -1:
        return None
    start = i + len(start_m)
    if end_m is None:
        chunk = text[start:]
    else:
        j = text.find(end_m, start)
        if j == -1:
            chunk = text[start:]
        else:
            chunk = text[start:j]
    return chunk.strip()


def extract_blocks(body: str) -> dict[str, Optional[str]]:
    out: dict[str, Optional[str]] = {}
    for key, start_m, end_m in MARKERS:
        if key == "biography":
            out[key] = _slice_between(body, start_m, None)
        else:
            out[key] = _slice_between(body, start_m, end_m)
    return out


def _slice_up_to_markers(text: str, start_m: str, end_markers: tuple[str, ...]) -> Optional[str]:
    i = text.find(start_m)
    if i == -1:
        return None
    pos = i + len(start_m)
    end_pos = len(text)
    for em in end_markers:
        j = text.find(em, pos)
        if j != -1 and j < end_pos:
            end_pos = j
    return text[pos:end_pos].strip()


def extract_blocks_flexible(body: str, parse_cfg: dict[str, Any]) -> dict[str, Optional[str]]:
    """事務局フォーマットの揺れ（講義紹介／講師紹介）に対応。"""
    out: dict[str, Optional[str]] = {}

    out["target_grade"] = _slice_up_to_markers(body, "【全学年対象】", ("【タイトル】",))
    out["title"] = _slice_up_to_markers(body, "【タイトル】", ("【講師】",))

    sp_text = _slice_up_to_markers(body, "【講師】", SPEAKER_END_MARKERS)
    out["speaker_block"] = sp_text

    desc = None
    for sm in DESC_START_MARKERS:
        chunk = _slice_up_to_markers(body, sm, tuple(BIO_START_MARKERS))
        if chunk:
            desc = chunk
            break
    out["description"] = desc

    bio = ""
    for bm in BIO_START_MARKERS:
        bm = bm.strip()
        k = body.find(bm)
        if k != -1:
            chunk = body[k + len(bm) :].strip()
            for sep in ("\n\nハワイ大学", "\nハワイ大学", "\n----------------------------------------------------------------", "\n──"):
                j = chunk.find(sep)
                if j != -1:
                    chunk = chunk[:j].strip()
            bio = chunk
            break
    out["biography"] = bio or None

    if out.get("title") is None and body.find("【タイトル】") == -1:
        return extract_blocks(body)
    return out


def parse_date_from_text(text: str) -> Optional[date]:
    if not text:
        return None
    patterns = [
        r"(?P<y>\d{4})\s*年\s*(?P<m>\d{1,2})\s*月\s*(?P<d>\d{1,2})\s*日",
        r"(?P<y>\d{4})\s*[/／]\s*(?P<m>\d{1,2})\s*[/／]\s*(?P<d>\d{1,2})",
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            try:
                return date(int(m.group("y")), int(m.group("m")), int(m.group("d")))
            except ValueError:
                continue
    return None


def split_body_into_sessions(body: str, enabled: bool) -> list[str]:
    if not enabled:
        return [body]
    n_t = body.count("【タイトル】")
    if n_t < 2:
        return [body]
    segs = re.split(r"(?=【タイトル】)", body)
    segs = [s.strip() for s in segs if s.strip()]
    ok: list[str] = []
    for s in segs:
        if "【講師】" in s and ("【タイトル】" in s or s.startswith("【タイトル】")):
            ok.append(s)
    return ok if ok else [body]


def normalize_token(s: str) -> str:
    """講義の重複判定用（記号・空白を弱める）。"""
    if not s:
        return ""
    t = str(s).replace("　", " ").strip().lower()
    t = re.sub(r"\s+", " ", t)
    t = re.sub(r"\s*先生\s*$", "", t)
    t = re.sub(r'[「」『』"\']', "", t)
    return t.strip()


def cell_to_date(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val)[:10]
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def parse_subject_date(
    subject: str, extra_patterns: Optional[list[str]] = None
) -> Optional[date]:
    patterns = list(DEFAULT_SUBJECT_DATE_PATTERNS)
    if extra_patterns:
        patterns.extend(p for p in extra_patterns if p)
    for pat in patterns:
        m = re.search(pat, subject)
        if m:
            y, mo, d = int(m.group("y")), int(m.group("m")), int(m.group("d"))
            try:
                return date(y, mo, d)
            except ValueError:
                continue
    return None


def split_speaker_block(block: Optional[str]) -> tuple[str, str]:
    if not block:
        return "", ""
    lines = [ln.strip() for ln in block.splitlines() if ln.strip()]
    if not lines:
        return "", ""
    name = lines[0]
    affiliation = "\n".join(lines[1:]).strip()
    return name, affiliation


@dataclass
class ParsedRow:
    lecture_date: date
    target_grade: str
    title: str
    speaker_name: str
    speaker_affiliation: str
    description: str
    biography: str
    email_subject: str
    email_received_at: datetime


@dataclass
class FailedRow:
    email_subject: str
    email_received_at: datetime
    reason: str
    body_preview: str = ""


def parse_mail_sessions(
    subject: str,
    body: str,
    received_at: datetime,
    parse_cfg: dict[str, Any],
) -> tuple[list[ParsedRow], list[FailedRow]]:
    extra_subject = (parse_cfg or {}).get("subject_date_patterns") or []
    body_scan = body[:200000] if body else ""
    base_dt = (
        parse_subject_date(subject, extra_subject)
        or parse_date_from_text(subject)
        or parse_date_from_text(body_scan)
    )

    split_on = parse_cfg.get("split_multiple_sessions", True)
    sessions = split_body_into_sessions(body, bool(split_on))
    allow_empty_bio = bool(parse_cfg.get("allow_empty_biography", True))
    allow_empty_desc = bool(parse_cfg.get("allow_empty_description", True))

    rows: list[ParsedRow] = []
    failed: list[FailedRow] = []

    for idx, seg in enumerate(sessions):
        blocks = extract_blocks_flexible(seg, parse_cfg)
        if not (blocks.get("title") or "").strip():
            fb = extract_blocks(seg)
            if (fb.get("title") or "").strip():
                blocks = fb

        ld = parse_date_from_text(seg) or base_dt
        if not ld:
            failed.append(
                FailedRow(
                    subject,
                    received_at,
                    "件名・本文セグメントから開催日を抽出できませんでした",
                    seg[:800],
                )
            )
            continue

        target_grade = (blocks.get("target_grade") or "").strip()
        title = (blocks.get("title") or "").strip()
        sp_name, sp_aff = split_speaker_block(blocks.get("speaker_block"))
        description = (blocks.get("description") or "").strip()
        biography = (blocks.get("biography") or "").strip()

        missing: list[str] = []
        if not title:
            missing.append("【タイトル】")
        if not sp_name:
            missing.append("【講師】")
        if not description and not allow_empty_desc:
            missing.append("講義紹介（【講義紹介文】等）")
        if not biography and not allow_empty_bio:
            missing.append("【講師略歴】等")

        if missing:
            failed.append(
                FailedRow(
                    subject,
                    received_at,
                    f"本文ブロックが不足: {', '.join(missing)}",
                    seg[:800],
                )
            )
            continue

        if not biography and allow_empty_bio:
            biography = "（本文に略歴ブロックなし）"
        if not description and allow_empty_desc:
            description = "（本文に講義紹介ブロックなし）"

        subj_out = subject.strip()
        if len(sessions) > 1:
            subj_out = f"{subj_out} [#session{idx + 1}]"

        rows.append(
            ParsedRow(
                lecture_date=ld,
                target_grade=target_grade or "全学年対象",
                title=title,
                speaker_name=sp_name,
                speaker_affiliation=sp_aff,
                description=description,
                biography=biography,
                email_subject=subj_out,
                email_received_at=received_at,
            )
        )

    return rows, failed


def parse_mail(
    subject: str,
    body: str,
    received_at: datetime,
    parse_cfg: dict[str, Any],
) -> tuple[Optional[ParsedRow], Optional[FailedRow]]:
    """単一セッション想定の互換 API（フィクスチャ用）。"""
    rows, fails = parse_mail_sessions(subject, body, received_at, parse_cfg)
    if rows:
        return rows[0], None
    if fails:
        return None, fails[0]
    return None, None


# ---------------------------------------------------------------------------
# Outlook
# ---------------------------------------------------------------------------

def _com_received_to_datetime(rv: Any) -> datetime:
    if isinstance(rv, datetime):
        return rv.replace(tzinfo=None) if rv.tzinfo else rv
    try:
        ts = rv.timestamp()  # pywintypes.Time
        return datetime.fromtimestamp(int(ts))
    except Exception:
        s = str(rv)
        for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(s[: len(fmt) + 2], fmt)
            except ValueError:
                continue
    return datetime.now()


def iter_subfolders(folder: Any) -> Iterator[Any]:
    try:
        subs = folder.Folders
    except Exception:
        return
    for i in range(1, int(subs.Count) + 1):
        yield subs.Item(i)


def find_folder_by_name(root_folder: Any, name: str) -> Optional[Any]:
    if getattr(root_folder, "Name", "") == name:
        return root_folder
    for sub in iter_subfolders(root_folder):
        found = find_folder_by_name(sub, name)
        if found is not None:
            return found
    return None


def get_outlook_folder(folder_name: str) -> Any:
    try:
        import win32com.client  # type: ignore
    except ImportError as e:
        raise RuntimeError("pywin32 is required for Outlook.") from e

    namespace = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    name = (folder_name or "").strip()
    # 空・Inbox 指定は既定の受信トレイをそのまま使う（サブフォルダを置かない運用向け）
    if not name or name.lower() in ("inbox", "受信トレイ", "__inbox__"):
        return namespace.GetDefaultFolder(6)

    inbox = namespace.GetDefaultFolder(6)
    found = find_folder_by_name(inbox, name)
    if found is not None:
        return found
    stores = namespace.Folders
    for i in range(1, int(stores.Count) + 1):
        store_root = stores.Item(i)
        found = find_folder_by_name(store_root, name)
        if found is not None:
            return found
    raise FileNotFoundError(
        f'フォルダ "{name}" が見つかりません。'
        "受信トレイ配下に同名フォルダを作成するか、config の outlook.folder_name を空にして受信トレイを使ってください。"
    )


def in_date_range(dt: datetime, start_d: date, end_d: date) -> bool:
    day = dt.date()
    return start_d <= day <= end_d


def normalize_smtp(addr: str) -> str:
    return (addr or "").strip().lower()


def get_sender_smtp(msg: Any) -> str:
    """SMTP アドレスを取得（インターネットメール／Exchange 双方を試す）。"""
    try:
        s = str(getattr(msg, "SenderEmailAddress", "") or "")
        if "@" in s and not s.strip().startswith("/"):
            return normalize_smtp(s)
    except Exception:
        pass
    try:
        ae = getattr(msg, "Sender", None)
        if ae is not None:
            ex = ae.GetExchangeUser()
            if ex is not None:
                s = str(getattr(ex, "PrimarySmtpAddress", "") or "")
                if "@" in s:
                    return normalize_smtp(s)
    except Exception:
        pass
    return ""


def message_matches_filters(
    msg: Any,
    subject: str,
    sender_email: Optional[str],
    subject_must_contain: Optional[str],
) -> bool:
    """送信元・件名キーワード（設定時のみ）。空の設定はフィルタしない。"""
    if subject_must_contain and subject_must_contain.strip():
        needle = subject_must_contain.strip()
        if needle.upper() not in subject.upper():
            return False
    if sender_email and sender_email.strip():
        want = normalize_smtp(sender_email)
        got = get_sender_smtp(msg)
        if not got or got != want:
            return False
    return True


def iter_mail_messages(folder: Any, start_d: date, end_d: date) -> Iterator[Any]:
    items = folder.Items
    try:
        items.Sort("[ReceivedTime]", True)
    except Exception:
        pass
    for i in range(1, int(items.Count) + 1):
        try:
            it = items.Item(i)
        except Exception:
            continue
        if getattr(it, "Class", None) != 43:
            continue
        try:
            rv = it.ReceivedTime
        except Exception:
            continue
        rdt = _com_received_to_datetime(rv)
        if in_date_range(rdt, start_d, end_d):
            yield it


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

COLUMNS: tuple[str, ...] = (
    "lecture_date",
    "target_grade",
    "title",
    "speaker_name",
    "speaker_affiliation",
    "description",
    "biography",
    "email_subject",
    "email_received_at",
)


def lecture_dedupe_key(d: dict[str, Any]) -> tuple[Optional[date], str, str]:
    cd = cell_to_date(d.get("lecture_date"))
    return (
        cd,
        normalize_token(str(d.get("speaker_name") or "")),
        normalize_token(str(d.get("title") or "")),
    )


def row_dict_from_parsed(p: ParsedRow) -> dict[str, Any]:
    return {
        "lecture_date": p.lecture_date,
        "target_grade": p.target_grade,
        "title": p.title,
        "speaker_name": p.speaker_name,
        "speaker_affiliation": p.speaker_affiliation,
        "description": p.description,
        "biography": p.biography,
        "email_subject": p.email_subject,
        "email_received_at": p.email_received_at,
    }


def load_lectures_workbook_rows(xlsx_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    if not xlsx_path.is_file():
        return list(COLUMNS), []
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "lectures" not in wb.sheetnames:
        wb.close()
        return list(COLUMNS), []
    ws = wb["lectures"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [str(h) if h is not None else "" for h in headers]
    rows_out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        d: dict[str, Any] = {}
        for i, hn in enumerate(headers):
            if not hn:
                continue
            d[hn] = row[i] if i < len(row) else None
        rows_out.append(d)
    wb.close()
    return headers, rows_out


def dedupe_failed_rows(failed: list[FailedRow]) -> list[FailedRow]:
    """同一件名・同一理由の失敗は、受信が新しい1行にまとめる（リマインド重複を整理）。"""
    best: dict[tuple[str, str], FailedRow] = {}
    for f in failed:
        key = ((f.email_subject or "").strip(), (f.reason or "").strip())
        cur = best.get(key)
        if cur is None or f.email_received_at > cur.email_received_at:
            best[key] = f
    return sorted(
        best.values(),
        key=lambda x: x.email_received_at,
        reverse=True,
    )


def merge_and_dedupe_rows(
    existing: list[dict[str, Any]],
    parsed_new: list[ParsedRow],
) -> list[dict[str, Any]]:
    combined = existing + [row_dict_from_parsed(p) for p in parsed_new]
    _drop_legacy = ("lecture_time_start", "lecture_time_end")
    for d in combined:
        for k in _drop_legacy:
            d.pop(k, None)
    combined.sort(
        key=lambda d: d.get("email_received_at") or datetime.min,
        reverse=True,
    )
    seen: set[tuple[Optional[date], str, str]] = set()
    kept: list[dict[str, Any]] = []
    for d in combined:
        k = lecture_dedupe_key(d)
        if k[0] is None:
            continue
        if k in seen:
            continue
        seen.add(k)
        kept.append(d)
    kept.sort(
        key=lambda x: (
            cell_to_date(x.get("lecture_date")) or date.min,
            str(x.get("title") or ""),
        )
    )
    return kept


def build_column_order(
    header_pref: list[str], rows: list[dict[str, Any]]
) -> list[str]:
    h: list[str] = []
    for c in COLUMNS:
        if c not in h:
            h.append(c)
    for x in header_pref:
        if x and x not in h:
            h.append(x)
    for r in rows:
        for k in r:
            if k and k not in h:
                h.append(k)
    return h


def write_lectures_workbook(
    xlsx_path: Path,
    header_pref: list[str],
    rows: list[dict[str, Any]],
    failed: list[FailedRow],
) -> None:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    headers = build_column_order(header_pref, rows)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "lectures"
    ws.append(headers)
    for r in rows:
        ws.append([r.get(col) for col in headers])
    fws = wb.create_sheet("failed")
    fws.append(["email_subject", "email_received_at", "reason", "body_preview"])
    for f in failed:
        fws.append([f.email_subject, f.email_received_at, f.reason, f.body_preview])
    wb.save(xlsx_path)
    wb.close()


def finalize_excel(
    xlsx_path: Path,
    new_rows: list[ParsedRow],
    failed: list[FailedRow],
) -> None:
    """既存台帳とマージし、開催日+講師+タイトルで重複除去して保存。"""
    h_old, existing = load_lectures_workbook_rows(xlsx_path)
    merged = merge_and_dedupe_rows(existing, new_rows)
    pref = h_old if h_old else list(COLUMNS)
    write_lectures_workbook(xlsx_path, pref, merged, dedupe_failed_rows(failed))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def setup_logging(log_dir: Path) -> None:
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / f"extract_{datetime.now().strftime('%Y-%m-%d')}.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )


def run_fixture(fixture_path: Path, cfg: dict[str, Any], root: Path) -> int:
    del root  # reserved
    data = json.loads(fixture_path.read_text(encoding="utf-8"))
    subject = data.get("subject", "")
    body = data.get("body", "")
    ra_s = data.get("email_received_at", "2026-05-01T09:00:00")
    received_at = datetime.fromisoformat(ra_s)
    parse_cfg = cfg.get("parse") or {}
    ok, fail = parse_mail(subject, body, received_at, parse_cfg)
    print(json.dumps(asdict(ok) if ok else None, default=str, ensure_ascii=False, indent=2))
    if fail:
        print("--- FAILED ---")
        print(json.dumps(asdict(fail), default=str, ensure_ascii=False, indent=2))
        return 1
    return 0


def run_extract(config_path: Path, dry_run: bool) -> int:
    cfg = load_config(config_path)
    root = config_path.parent
    parse_cfg = cfg.get("parse") or {}

    outlook_cfg = cfg.get("outlook") or {}
    folder_name = outlook_cfg.get("folder_name")
    folder_name = "" if folder_name is None else str(folder_name)

    recv_from_s = outlook_cfg.get("received_date_from") or "2024-06-01"
    recv_to_s = outlook_cfg.get("received_date_to") or "2026-12-31"
    recv_start = date.fromisoformat(str(recv_from_s)[:10])
    recv_end = date.fromisoformat(str(recv_to_s)[:10])

    lec_from_s = outlook_cfg.get("lecture_date_from") or "2025-01-01"
    lec_to_s = outlook_cfg.get("lecture_date_to") or "2026-05-31"
    lec_from = date.fromisoformat(str(lec_from_s)[:10])
    lec_to = date.fromisoformat(str(lec_to_s)[:10])

    paths = cfg.get("paths") or {}
    lectures_rel = paths.get("lectures_xlsx") or "data/lectures.xlsx"
    logs_rel = paths.get("logs_dir") or "logs"
    xlsx_path = resolve_under_root(root, lectures_rel)
    log_dir = resolve_under_root(root, logs_rel)

    setup_logging(log_dir)

    new_parsed: list[ParsedRow] = []
    failed_all: list[FailedRow] = []

    sender_email = (outlook_cfg.get("sender_email") or "").strip() or None
    subject_must_contain = (outlook_cfg.get("subject_must_contain") or "").strip() or None

    folder = get_outlook_folder(folder_name)
    for msg in iter_mail_messages(folder, recv_start, recv_end):
        try:
            subj = str(getattr(msg, "Subject", "") or "")
            body = str(getattr(msg, "Body", "") or "")
            rdt = _com_received_to_datetime(msg.ReceivedTime)
        except Exception as e:
            logging.warning("skip mail: %s", e)
            continue

        if not message_matches_filters(msg, subj, sender_email, subject_must_contain):
            logging.info(
                "skip (sender/subject filter): %s",
                subj[:80],
            )
            continue

        parsed_list, fail_list = parse_mail_sessions(subj, body, rdt, parse_cfg)
        for f in fail_list:
            logging.info("parse failed: %s — %s", subj[:80], f.reason)
        failed_all.extend(fail_list)
        for ok in parsed_list:
            if ok.lecture_date < lec_from or ok.lecture_date > lec_to:
                continue
            new_parsed.append(ok)

    if dry_run:
        logging.info("dry-run: %d new rows, %d failed", len(new_parsed), len(failed_all))
        for r in new_parsed:
            logging.info("  NEW: %s %s", r.lecture_date, r.title[:60])
        return 0

    try:
        finalize_excel(xlsx_path, new_parsed, failed_all)
    except ValueError as e:
        logging.error("%s", e)
        return 5

    logging.info("done: added %d rows, failed %d", len(new_parsed), len(failed_all))
    return 0


def main(argv: Optional[list[str]] = None) -> int:
    p = argparse.ArgumentParser(description="HMEP: extract lectures from Outlook to Excel")
    p.add_argument(
        "--config",
        type=Path,
        default=PIPELINE_ROOT / "config.yaml",
        help="config.yaml path (default: hmep_pipeline/config.yaml)",
    )
    p.add_argument("--dry-run", action="store_true", help="Excel を書かずログのみ")
    p.add_argument(
        "--fixture",
        type=Path,
        help="JSON フィクスチャ (subject, body, email_received_at) でパースのみ検証",
    )
    args = p.parse_args(argv)

    cfg_path = args.config.resolve()
    if not cfg_path.is_file():
        print(f"config not found: {cfg_path}", file=sys.stderr)
        return 2

    cfg = load_config(cfg_path)
    if args.fixture is not None:
        return run_fixture(args.fixture.resolve(), cfg, cfg_path.parent)

    try:
        return run_extract(cfg_path, args.dry_run)
    except FileNotFoundError as e:
        print(e, file=sys.stderr)
        return 3
    except RuntimeError as e:
        print(e, file=sys.stderr)
        return 4


if __name__ == "__main__":
    raise SystemExit(main())

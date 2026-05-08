"""
② 動画リネーム: lectures.xlsx と動画フォルダを突合し、videos_renamed/ にコピーする。

マッチング（優先順）:
1) 動画の更新時刻が開催日時の±time_window_hours以内 かつ 類似度≥similarity_threshold
2) 未使用動画のうちパスに開催日が含まれるもので類似度最大が≥similarity_fallback
3) 時間窓内の動画で類似度最大が≥similarity_fallback（開催日近傍の収録を優先）
使用例（hmep_pipeline/02_rename から）:
  python rename_videos.py --dry-run
  python rename_videos.py
"""

from __future__ import annotations

import argparse
import logging
import re
import shutil
import sys
from datetime import date, datetime, time
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Optional

import yaml
from openpyxl import load_workbook

PIPELINE_ROOT = Path(__file__).resolve().parent.parent

RENAME_COLS: tuple[str, ...] = (
    "source_video_file",
    "renamed_video_file",
    "rename_match_score",
    "rename_status",
)

WIN_ILLEGAL = '<>:"/\\|?*'


def load_config(path: Path) -> dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def resolve_path(root: Path, rel: str) -> Path:
    p = Path(rel)
    if p.is_absolute():
        return p.resolve()
    return (root / rel).resolve()


def sanitize_component(s: str) -> str:
    s = s or ""
    out = "".join(c if c not in WIN_ILLEGAL else " " for c in s)
    return re.sub(r"\s+", " ", out).strip()


def build_target_filename(
    lecture_date: date,
    speaker_name: str,
    title: str,
    ext: str,
) -> str:
    y, m, d = lecture_date.year, lecture_date.month, lecture_date.day
    sp = sanitize_component(speaker_name)
    if not sp.endswith("先生"):
        sp = f"{sp} 先生"
    tit = sanitize_component(title).replace('"', "'")
    # Windows はファイル名に " が使えないため、タイトルを「」で囲む
    base = f"{y} {m} {d} {sp} 「{tit}」"
    base = sanitize_component(base)
    e = ext if ext.startswith(".") else f".{ext}"
    return base + e.lower()


def parse_hhmm(s: Any) -> Optional[time]:
    if s is None:
        return None
    if isinstance(s, time):
        t = s
        return time(t.hour, t.minute, t.second if t.second else 0)
    if isinstance(s, datetime):
        return time(s.hour, s.minute, s.second)
    m = re.match(r"(\d{1,2}):(\d{2})", str(s).strip())
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    try:
        return time(h, mi)
    except ValueError:
        return None


def to_date(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    try:
        s = str(val)[:10]
        return date.fromisoformat(s)
    except (ValueError, TypeError):
        return None


def lecture_datetime(ld: date, start_s: Any) -> datetime:
    t = parse_hhmm(start_s)
    if t is None:
        t = time(12, 0)
    return datetime.combine(ld, t)


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").lower()).strip()


def similarity_to_filename(path: Path, speaker: str, title: str) -> float:
    stem = path.stem
    a = norm(stem)
    b = norm(speaker) + " " + norm(title)
    return SequenceMatcher(None, a, b).ratio()


def within_hours(a: datetime, b: datetime, hours: float) -> bool:
    return abs((a - b).total_seconds()) <= hours * 3600.0


def mtime_dt(p: Path) -> datetime:
    return datetime.fromtimestamp(p.stat().st_mtime)


def path_suggests_lecture_date(vp: Path, ld: date) -> bool:
    """パスまたはファイル名に、台帳の開催日と解釈できる日付が含まれるか（緩いヒント）。"""
    s = str(vp.resolve()).replace("\\", "/")
    y, m, d = ld.year, ld.month, ld.day
    needle_variants = [
        f"{y}-{m:02d}-{d:02d}",
        f"{y}-{m}-{d}",
        f"{y}.{m}.{d}",
        f"{y}.{m:02d}.{d:02d}",
        f"{y}/{m:02d}/{d:02d}",
        f"{y}年{m}月{d}日",
        f"{y}{m:02d}{d:02d}",
    ]
    return any(n in s for n in needle_variants)


def select_video_for_row(
    all_videos: list[Path],
    used: set[Path],
    speaker: str,
    title: str,
    ld: date,
    ldt: datetime,
    window_h: float,
    threshold: float,
    fallback: float,
) -> tuple[Optional[Path], float, str]:
    """候補動画1本を選ぶ。

    - strict: 時間窓内かつ類似度>=threshold
    - relaxed_date: パスに開催日が含まれる未使用動画のうち類似度最大が>=fallback
    - relaxed_window: 時間窓内の類似度最大が>=fallback（厳密阈未満でも採用。誤割当は global より抑えやすい）
    """
    best_w: Optional[Path] = None
    best_w_sc = -1.0
    for vp in all_videos:
        if vp in used:
            continue
        if not within_hours(mtime_dt(vp), ldt, window_h):
            continue
        sc = similarity_to_filename(vp, speaker, title)
        if sc > best_w_sc:
            best_w_sc = sc
            best_w = vp
    if best_w is not None and best_w_sc >= threshold:
        return best_w, best_w_sc, "strict"

    best_d: Optional[Path] = None
    best_d_sc = -1.0
    for vp in all_videos:
        if vp in used:
            continue
        if not path_suggests_lecture_date(vp, ld):
            continue
        sc = similarity_to_filename(vp, speaker, title)
        if sc > best_d_sc:
            best_d_sc = sc
            best_d = vp
    if best_d is not None and best_d_sc >= fallback:
        return best_d, best_d_sc, "relaxed_date"

    if best_w is not None and best_w_sc >= fallback:
        return best_w, best_w_sc, "relaxed_window"

    report_sc = max(best_w_sc, best_d_sc)
    return None, report_sc, "none"


def collect_videos(root: Path, exts: set[str], recursive: bool) -> list[Path]:
    out: list[Path] = []
    if not root.is_dir():
        return out
    if recursive:
        for p in root.rglob("*"):
            if p.is_file() and p.suffix.lower() in exts:
                out.append(p)
    else:
        for p in root.iterdir():
            if p.is_file() and p.suffix.lower() in exts:
                out.append(p)
    return out


def load_sheet(
    xlsx: Path,
) -> tuple[list[str], list[dict[str, Any]]]:
    if not xlsx.is_file():
        return [], []
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    if "lectures" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"{xlsx} に lectures シートがありません")
    ws = wb["lectures"]
    headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        if all(v is None for v in row):
            continue
        d: dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            d[h] = row[i] if i < len(row) else None
        rows.append(d)
    wb.close()
    return headers, rows


def save_lectures_sheet(
    xlsx: Path,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> None:
    wb = load_workbook(xlsx)
    if "lectures" not in wb.sheetnames:
        wb.close()
        raise ValueError("lectures sheet missing")
    ws = wb["lectures"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h or None)
    for ri, rd in enumerate(rows, start=2):
        for c, h in enumerate(headers, start=1):
            if not h:
                continue
            ws.cell(row=ri, column=c, value=rd.get(h))
    wb.save(xlsx)
    wb.close()


def setup_logging(log_dir: Path) -> None:
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / f"rename_{datetime.now().strftime('%Y-%m-%d')}.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )


def pick_unique_dest(dest_dir: Path, filename: str) -> Path:
    dest = dest_dir / filename
    if not dest.exists():
        return dest
    stem = Path(filename).stem
    suf = Path(filename).suffix
    n = 2
    while True:
        cand = dest_dir / f"{stem}_{n}{suf}"
        if not cand.exists():
            return cand
        n += 1


def run_rename(
    config_path: Path,
    dry_run: bool,
    force: bool,
    limit: Optional[int] = None,
    time_window_hours_override: Optional[float] = None,
    similarity_threshold_override: Optional[float] = None,
    similarity_fallback_override: Optional[float] = None,
) -> int:
    cfg = load_config(config_path)
    root = config_path.parent

    r_cfg = cfg.get("rename") or {}
    outlook = cfg.get("outlook") or {}
    date_from_s = r_cfg.get("date_from") or outlook.get("date_from") or "2025-01-01"
    date_to_s = r_cfg.get("date_to") or outlook.get("date_to") or "2026-05-31"
    start_d = date.fromisoformat(str(date_from_s)[:10])
    end_d = date.fromisoformat(str(date_to_s)[:10])

    window_h = float(r_cfg.get("time_window_hours", 24))
    if time_window_hours_override is not None:
        window_h = float(time_window_hours_override)
    threshold = float(r_cfg.get("similarity_threshold", 0.22))
    if similarity_threshold_override is not None:
        threshold = float(similarity_threshold_override)
    fallback = float(r_cfg.get("similarity_fallback", 0.10))
    if similarity_fallback_override is not None:
        fallback = float(similarity_fallback_override)
    recursive = bool(r_cfg.get("recursive", True))
    exts = {str(x).lower() for x in (r_cfg.get("extensions") or [".mp4"])}

    paths = cfg.get("paths") or {}
    videos_rel = paths.get("videos_dir") or "videos"
    out_rel = paths.get("videos_renamed_dir") or "videos_renamed"
    lectures_rel = paths.get("lectures_xlsx") or "data/lectures.xlsx"
    logs_rel = paths.get("logs_dir") or "logs"

    videos_dir = resolve_path(root, videos_rel)
    dest_dir = resolve_path(root, out_rel)
    xlsx_path = resolve_path(root, lectures_rel)
    log_dir = resolve_path(root, logs_rel)

    setup_logging(log_dir)

    if not xlsx_path.is_file():
        logging.error("lectures.xlsx がありません: %s", xlsx_path)
        return 2

    headers, rows = load_sheet(xlsx_path)
    for c in RENAME_COLS:
        if c not in headers:
            headers.append(c)

    all_videos = collect_videos(videos_dir, exts, recursive)
    logging.info(
        "videos: %d files under %s (recursive=%s)",
        len(all_videos),
        videos_dir,
        recursive,
    )
    logging.info("time_window_hours=%s", window_h)
    logging.info(
        "similarity threshold=%s fallback (path or window)=%s",
        threshold,
        fallback,
    )

    used: set[Path] = set()
    max_renames = limit if limit is not None and limit > 0 else 0
    n_renamed = 0

    for row in rows:
        ld = to_date(row.get("lecture_date"))
        if ld is None:
            continue
        if ld < start_d or ld > end_d:
            continue

        if not force and row.get("renamed_video_file"):
            row.setdefault("rename_status", "skipped_already_renamed")
            continue

        speaker = str(row.get("speaker_name") or "").strip()
        title = str(row.get("title") or "").strip()
        # 開始時刻列は廃止。マッチングの基準時刻は開催日の正午（±24h 窓）
        ldt = lecture_datetime(ld, None)

        best, best_sc, how = select_video_for_row(
            all_videos,
            used,
            speaker,
            title,
            ld,
            ldt,
            window_h,
            threshold,
            fallback,
        )
        if how == "none" or best is None:
            row["rename_status"] = "skipped_no_match"
            row["rename_match_score"] = round(best_sc, 4) if best_sc >= 0 else None
            logging.info(
                "no match: %s %s | best_score=%.3f",
                ld,
                title[:40],
                best_sc,
            )
            continue

        if max_renames and n_renamed >= max_renames:
            row["rename_status"] = "skipped_limit"
            continue

        assert best is not None
        status_for_tag = {
            "strict": "ok",
            "relaxed_date": "ok_relaxed_date",
            "relaxed_window": "ok_relaxed_window",
        }

        fname = build_target_filename(ld, speaker, title, best.suffix)
        dest_path = pick_unique_dest(dest_dir, fname)

        try:
            rel_src = str(best.relative_to(videos_dir.resolve()))
        except ValueError:
            rel_src = str(best)

        if dry_run:
            logging.info(
                "[dry-run] COPY %s -> %s (score=%.3f %s)",
                rel_src,
                dest_path.name,
                best_sc,
                how,
            )
            row["rename_status"] = f"dry_run_{status_for_tag.get(how, 'ok')}"
            row["source_video_file"] = rel_src
            row["renamed_video_file"] = dest_path.name
            row["rename_match_score"] = round(best_sc, 4)
            n_renamed += 1
            continue

        dest_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(best, dest_path)
        used.add(best)

        row["source_video_file"] = rel_src
        row["renamed_video_file"] = dest_path.name
        row["rename_match_score"] = round(best_sc, 4)
        row["rename_status"] = status_for_tag.get(how, "ok")
        logging.info(
            "copied: %s -> %s (score=%.3f %s)",
            rel_src,
            dest_path.name,
            best_sc,
            how,
        )
        n_renamed += 1

    if dry_run:
        logging.info("dry-run: no files copied, no xlsx save")
        return 0

    save_lectures_sheet(xlsx_path, headers, rows)
    logging.info("saved: %s", xlsx_path)
    return 0


def main(argv: Optional[list[str]] = None) -> int:
    p = argparse.ArgumentParser(description="HMEP: rename video files from lectures.xlsx")
    p.add_argument(
        "--config",
        type=Path,
        default=PIPELINE_ROOT / "config.yaml",
    )
    p.add_argument("--dry-run", action="store_true")
    p.add_argument(
        "--force",
        action="store_true",
        help="既に renamed_video_file がある行も再マッチ",
    )
    p.add_argument(
        "--limit",
        type=int,
        default=None,
        metavar="N",
        help="コピー成功させる最大件数（試験用。省略時は制限なし）",
    )
    p.add_argument(
        "--time-window-hours",
        type=float,
        default=None,
        metavar="H",
        help="マッチングの±時間窓（時間）。指定時は config の rename.time_window_hours より優先",
    )
    p.add_argument(
        "--similarity-threshold",
        type=float,
        default=None,
        metavar="T",
        help="厳密マッチの類似度下限（既定は config の rename.similarity_threshold）",
    )
    p.add_argument(
        "--similarity-fallback",
        type=float,
        default=None,
        metavar="T",
        help="フォールバックの類似度下限（rename.similarity_fallback）",
    )
    args = p.parse_args(argv)

    cfg_path = args.config.resolve()
    if not cfg_path.is_file():
        print(f"config not found: {cfg_path}", file=sys.stderr)
        return 2

    try:
        return run_rename(
            cfg_path,
            args.dry_run,
            args.force,
            limit=args.limit,
            time_window_hours_override=args.time_window_hours,
            similarity_threshold_override=args.similarity_threshold,
            similarity_fallback_override=args.similarity_fallback,
        )
    except ValueError as e:
        print(e, file=sys.stderr)
        return 3


if __name__ == "__main__":
    raise SystemExit(main())
